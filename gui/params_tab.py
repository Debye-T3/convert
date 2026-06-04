"""Tab 2: Batch parameter defaults + per-file override table (Excel-like)."""

from pathlib import Path

from PySide6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QFormLayout, QLineEdit,
    QLabel, QTableWidget, QTableWidgetItem, QHeaderView, QGroupBox,
    QPushButton, QSplitter, QComboBox, QMenu, QFileDialog, QMessageBox,
    QInputDialog,
)
from PySide6.QtCore import Qt
from PySide6.QtGui import QAction


STANDARD_KEYS = [
    "sample_name", "sample_id",
    "position_x", "position_y", "position_z",
    "position_polar", "position_tilt", "position_azimuth",
    "temperature_K", "photon_energy_eV", "polarization", "slit",
    "work_function_eV",
]

STANDARD_LABELS = [
    "Sample Name", "Sample ID",
    "X", "Y", "Z",
    "Polar", "Tilt", "Azimuth",
    "T (K)", "hv (eV)", "Polarization", "Slit",
    "WF Φ (eV)",
]

FIELD_DEFAULTS = {
    "work_function_eV": "4.2",
}

POLARIZATION_OPTIONS = ["", "p-pol", "s-pol", "circular", "linear", "none"]

CUSTOM_PREFIX = "user_"


class ParamsTab(QWidget):
    """Batch defaults on the left, per-file overrides on the right.

    Columns are drag-reorderable.  Custom parameter columns can be added /
    removed.  Excel import / template export are supported.
    """

    def __init__(self, parent=None):
        super().__init__(parent)
        self._file_paths = []
        self._overrides = {}          # fp -> {key: value}
        self._hidden_cols = set()     # logical parameter column indices

        # dynamic field lists (parallel arrays)
        self._field_keys = list(STANDARD_KEYS)
        self._field_labels = list(STANDARD_LABELS)

        # ---- build UI ----
        layout = QVBoxLayout(self)

        title = QLabel("Experiment Parameters")
        title.setStyleSheet("font-size: 14pt; font-weight: bold;")
        layout.addWidget(title)

        self.splitter = QSplitter(Qt.Horizontal)

        # left panel
        self._build_left_panel()
        # right panel
        self._build_right_panel()

        self.splitter.setSizes([360, 600])
        layout.addWidget(self.splitter)

        # nav
        nav = QHBoxLayout()
        nav.addWidget(QPushButton("← Back: Select Files", clicked=self._go_back))
        nav.addStretch()
        nav.addWidget(QPushButton("Next: Convert →", clicked=self._go_next))
        layout.addLayout(nav)

    # ================================================================
    #  left panel — batch defaults
    # ================================================================

    def _build_left_panel(self):
        left = QWidget()
        left_layout = QVBoxLayout(left)
        left_layout.setContentsMargins(0, 0, 0, 0)

        form_group = QGroupBox("Batch Defaults")
        self._form_layout = QFormLayout(form_group)
        self._field_widgets = {}
        self._populate_form()
        left_layout.addWidget(form_group)

        # custom fields section
        custom_group = QGroupBox("Custom Fields")
        custom_outer = QVBoxLayout(custom_group)
        self._custom_container = QVBoxLayout()
        self._custom_widgets = {}  # key -> (QLineEdit, QPushButton, QHBoxLayout)
        custom_outer.addLayout(self._custom_container)
        add_btn = QPushButton("+ Add Custom Field")
        add_btn.clicked.connect(self._add_custom_field)
        custom_outer.addWidget(add_btn)
        left_layout.addWidget(custom_group)

        tip = QLabel(
            "Default values apply to all files.\n"
            "Override per-file in the table →\n\n"
            "Right-click table headers to hide columns.\n"
            "Drag columns to reorder."
        )
        tip.setStyleSheet("color: #888; font-size: 10pt; padding: 8px;")
        left_layout.addWidget(tip)
        left_layout.addStretch()

        self.splitter.addWidget(left)

    def _populate_form(self):
        """(Re)build the standard-fields portion of the form."""
        # clear existing rows
        while self._form_layout.rowCount() > 0:
            self._form_layout.removeRow(0)
        self._field_widgets.clear()

        for key, label in zip(STANDARD_KEYS, STANDARD_LABELS):
            if key == "polarization":
                w = QComboBox()
                w.setEditable(True)
                w.addItems(POLARIZATION_OPTIONS)
                w.setCurrentText("")
            else:
                w = QLineEdit()
                w.setPlaceholderText(label)
                if key in FIELD_DEFAULTS:
                    w.setText(FIELD_DEFAULTS[key])
            self._form_layout.addRow(label, w)
            self._field_widgets[key] = w

    # ================================================================
    #  right panel — per-file table
    # ================================================================

    def _build_right_panel(self):
        right = QWidget()
        right_layout = QVBoxLayout(right)
        right_layout.setContentsMargins(0, 0, 0, 0)

        table_group = QGroupBox("Per-File Overrides")
        table_layout = QVBoxLayout(table_group)

        current_row = QHBoxLayout()
        current_file_title = QLabel("Current file:")
        current_file_title.setStyleSheet("color: #cbd5e1;")
        current_row.addWidget(current_file_title)
        self.current_file_label = QLabel("-")
        self.current_file_label.setTextInteractionFlags(Qt.TextSelectableByMouse)
        self.current_file_label.setStyleSheet(
            "font-weight: bold; padding: 4px 8px; color: #f3f4f6; "
            "background-color: #2b2f3a; border: 1px solid #4b5563; "
            "border-radius: 4px;"
        )
        current_row.addWidget(self.current_file_label, 1)
        table_layout.addLayout(current_row)

        self.table = QTableWidget()
        self.table.horizontalHeader().setSectionsMovable(True)
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
        self.table.horizontalHeader().setStretchLastSection(True)
        self.table.horizontalHeader().setMinimumSectionSize(50)
        self.table.horizontalHeader().setContextMenuPolicy(Qt.CustomContextMenu)
        self.table.horizontalHeader().customContextMenuRequested.connect(
            self._header_context_menu
        )
        self.table.verticalHeader().setSectionResizeMode(QHeaderView.Interactive)
        self.table.verticalHeader().setMinimumSectionSize(20)
        self.table.verticalHeader().setDefaultSectionSize(24)
        self.table.setAlternatingRowColors(True)
        self.table.cellChanged.connect(self._on_cell_changed)
        self.table.itemSelectionChanged.connect(self._update_current_file_from_selection)
        self.table.currentCellChanged.connect(self._on_current_cell_changed)
        table_layout.addWidget(self.table)

        # buttons under table
        btn_row = QHBoxLayout()

        import_btn = QPushButton("Import from Excel...")
        import_btn.setToolTip(
            "Import per-file parameters from an Excel spreadsheet.\n"
            "A 'file' or 'path' column must contain source data paths."
        )
        import_btn.clicked.connect(self._import_excel)
        btn_row.addWidget(import_btn)

        export_btn = QPushButton("Export Template")
        export_btn.setToolTip(
            "Export a standardized .xlsx template with the current\n"
            "column layout for batch filling."
        )
        export_btn.clicked.connect(self._export_template)
        btn_row.addWidget(export_btn)

        btn_row.addStretch()
        show_all_btn = QPushButton("Show All Columns")
        show_all_btn.clicked.connect(self._show_all_columns)
        btn_row.addWidget(show_all_btn)
        table_layout.addLayout(btn_row)

        table_tip = QLabel(
            "Click any cell to override. Right-click header to hide. "
            "Drag header to reorder columns."
        )
        table_tip.setStyleSheet("color: #888; font-size: 10pt; padding: 4px;")
        table_layout.addWidget(table_tip)

        right_layout.addWidget(table_group)
        self.splitter.addWidget(right)

    # ================================================================
    #  public API
    # ================================================================

    def set_files(self, file_paths):
        self._file_paths = file_paths
        self._rebuild_table()

    def get_all_params(self):
        """Return batch defaults dict with '_overrides' key."""
        batch = {}
        # standard fields from form
        for key, w in self._field_widgets.items():
            text = (w.currentText().strip() if isinstance(w, QComboBox)
                    else w.text().strip())
            if text:
                try:
                    batch[key] = float(text)
                except ValueError:
                    batch[key] = text
        # custom fields from custom widgets
        for key, (edit, _, _) in self._custom_widgets.items():
            text = edit.text().strip()
            if text:
                try:
                    batch[key] = float(text)
                except ValueError:
                    batch[key] = text
        batch["_overrides"] = dict(self._overrides)
        return batch

    # ================================================================
    #  table management
    # ================================================================

    @property
    def _all_field_keys(self):
        return self._field_keys

    @property
    def _all_field_labels(self):
        return self._field_labels

    def _rebuild_table(self):
        self.table.blockSignals(True)
        keys = self._all_field_keys
        labels = self._all_field_labels
        n_cols = len(keys) + 1

        self.table.setColumnCount(n_cols)
        self.table.setHorizontalHeaderLabels(["File"] + list(labels))
        self.table.setRowCount(len(self._file_paths))

        for i, fp in enumerate(self._file_paths):
            name_item = QTableWidgetItem(Path(fp).name)
            name_item.setFlags(name_item.flags() & ~Qt.ItemIsEditable)
            name_item.setToolTip(fp)
            self.table.setItem(i, 0, name_item)
            for j, key in enumerate(keys):
                col = j + 1
                val = self._overrides.get(fp, {}).get(key, "")
                self.table.setItem(i, col, QTableWidgetItem(str(val) if val else ""))

        for col in range(1, self.table.columnCount()):
            self.table.setColumnHidden(col, col in self._hidden_cols)
        self.table.setColumnHidden(0, False)
        self.table.blockSignals(False)
        if self._file_paths and self.table.currentRow() < 0:
            self.table.setCurrentCell(0, 1 if self.table.columnCount() > 1 else 0)
        self._update_current_file_label(self.table.currentRow())

    def _on_cell_changed(self, row, col):
        if row >= len(self._file_paths) or col <= 0:
            return
        self._update_current_file_label(row)
        fp = self._file_paths[row]
        key_index = col - 1
        if key_index < 0 or key_index >= len(self._all_field_keys):
            return
        key = self._all_field_keys[key_index]
        item = self.table.item(row, col)
        text = item.text().strip() if item else ""
        if fp not in self._overrides:
            self._overrides[fp] = {}
        if text:
            self._overrides[fp][key] = text
        else:
            self._overrides[fp].pop(key, None)
            if not self._overrides[fp]:
                del self._overrides[fp]

    def _on_current_cell_changed(self, row, _col, _previous_row, _previous_col):
        self._update_current_file_label(row)

    def _update_current_file_from_selection(self):
        items = self.table.selectedItems()
        if items:
            self._update_current_file_label(items[0].row())

    def _update_current_file_label(self, row):
        if 0 <= row < len(self._file_paths):
            path = self._file_paths[row]
            self.current_file_label.setText(Path(path).name)
            self.current_file_label.setToolTip(path)
        else:
            self.current_file_label.setText("-")
            self.current_file_label.setToolTip("")

    # ================================================================
    #  column hide / show
    # ================================================================

    def _header_context_menu(self, pos):
        col = self.table.horizontalHeader().logicalIndexAt(pos)
        if col <= 0:
            return
        key_index = col - 1
        if key_index < 0 or key_index >= len(self._all_field_keys):
            return
        is_custom = self._all_field_keys[key_index].startswith(CUSTOM_PREFIX)

        menu = QMenu(self)
        label = self._all_field_labels[key_index]
        hide_action = QAction(f"Hide '{label}'", self)
        hide_action.triggered.connect(lambda: self._hide_column(col))
        menu.addAction(hide_action)

        if is_custom:
            menu.addSeparator()
            remove_action = QAction(f"Remove '{label}'", self)
            remove_action.triggered.connect(lambda: self._remove_custom_field(key_index))
            menu.addAction(remove_action)
        menu.exec(self.table.horizontalHeader().mapToGlobal(pos))

    def _hide_column(self, col):
        if col <= 0:
            return
        self._hidden_cols.add(col)
        self.table.setColumnHidden(col, True)

    def _show_all_columns(self):
        self._hidden_cols.clear()
        for col in range(1, self.table.columnCount()):
            self.table.setColumnHidden(col, False)

    # ================================================================
    #  custom fields
    # ================================================================

    def _add_custom_field(self):
        name, ok = QInputDialog.getText(
            self, "Add Custom Field",
            "Field name (e.g. 'K deposition time'):"
        )
        if not ok or not name.strip():
            return
        key = CUSTOM_PREFIX + name.strip().lower().replace(" ", "_")
        label = name.strip()

        # avoid duplicates
        if key in self._field_keys:
            QMessageBox.warning(self, "Duplicate", f"Field '{label}' already exists.")
            return

        self._field_keys.append(key)
        self._field_labels.append(label)

        # add widget row in custom container
        edit = QLineEdit()
        edit.setPlaceholderText(label)
        remove_btn = QPushButton("✕")
        remove_btn.setFixedWidth(28)
        remove_btn.setToolTip(f"Remove '{label}'")
        remove_btn.clicked.connect(lambda: self._remove_custom_field_by_key(key))
        row_layout = QHBoxLayout()
        row_layout.addWidget(edit, 1)
        row_layout.addWidget(remove_btn)
        self._custom_container.addLayout(row_layout)
        self._custom_widgets[key] = (edit, remove_btn, row_layout)

        self._rebuild_table()

    def _remove_custom_field_by_key(self, key):
        """Remove custom field identified by key string."""
        try:
            idx = self._field_keys.index(key)
        except ValueError:
            return
        self._remove_custom_field(idx)

    def _remove_custom_field(self, key_index):
        key = self._field_keys[key_index]
        if not key.startswith(CUSTOM_PREFIX):
            return

        # confirm
        label = self._field_labels[key_index]
        reply = QMessageBox.question(
            self, "Remove Field",
            f"Remove field '{label}' and all its data?",
            QMessageBox.Yes | QMessageBox.No, QMessageBox.No
        )
        if reply != QMessageBox.Yes:
            return

        # remove from lists
        self._field_keys.pop(key_index)
        self._field_labels.pop(key_index)

        # clean widget
        if key in self._custom_widgets:
            edit, btn, row = self._custom_widgets.pop(key)
            # remove widgets from layout
            row.removeWidget(edit)
            row.removeWidget(btn)
            edit.deleteLater()
            btn.deleteLater()
            # remove the row layout itself
            while row.count():
                item = row.takeAt(0)
            self._custom_container.removeItem(row)

        # clean overrides referencing this key
        for fp in list(self._overrides.keys()):
            self._overrides[fp].pop(key, None)
            if not self._overrides[fp]:
                del self._overrides[fp]

        self._rebuild_table()

    # ================================================================
    #  Excel import
    # ================================================================

    def _import_excel(self):
        path, _ = QFileDialog.getOpenFileName(
            self, "Import Excel Parameters", "",
            "Excel Files (*.xlsx);;All Files (*)"
        )
        if not path:
            return
        try:
            import openpyxl
        except ImportError:
            QMessageBox.warning(
                self, "Missing Package",
                "The 'openpyxl' package is required to read Excel files.\n"
                "Install it with:  pip install openpyxl"
            )
            return

        try:
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            ws = wb.active
            row_iter = ws.iter_rows(values_only=True)
            header_row = next(row_iter, None)
            if not header_row:
                QMessageBox.warning(self, "Empty File", "No data found in Excel file.")
                wb.close()
                return

            header = [str(c).strip() if c else "" for c in header_row]
            file_col = None
            for ci, h in enumerate(header):
                if h.lower() in {"file", "path", "filepath", "file_path"}:
                    file_col = ci
                    break
            if file_col is None:
                QMessageBox.warning(
                    self, "Missing Column",
                    "Excel must contain a 'file' or 'path' column."
                )
                wb.close()
                return

            col_to_key = {}
            added_columns = []
            for ci, h in enumerate(header):
                if ci == file_col or not h:
                    continue
                matched = False
                for key, label in zip(self._all_field_keys, self._all_field_labels):
                    if h == label or h.lower() == key.lower():
                        col_to_key[ci] = key
                        matched = True
                        break
                if not matched:
                    for key, label in zip(self._all_field_keys, self._all_field_labels):
                        if h.lower().split("(")[0].strip() == label.lower().split("(")[0].strip():
                            col_to_key[ci] = key
                            matched = True
                            break
                if not matched:
                    key = self._user_key_from_header(h)
                    if key not in self._field_keys:
                        self._field_keys.append(key)
                        self._field_labels.append(h)
                        added_columns.append((key, h))
                    col_to_key[ci] = key

            for key, label in added_columns:
                self._add_custom_widget_for_key(key, label)

            excel_dir = Path(path).resolve().parent
            imported = 0
            added_files = []
            pending_overrides = {}
            for row in row_iter:
                if not row or file_col >= len(row) or not row[file_col]:
                    continue
                src = self._resolve_excel_source(str(row[file_col]).strip(), excel_dir)
                if src is None:
                    continue
                fp = str(src)
                if fp not in self._file_paths and fp not in added_files:
                    added_files.append(fp)
                overrides = {}
                for ci, key in col_to_key.items():
                    val = row[ci] if ci < len(row) else None
                    if val is not None and str(val).strip():
                        overrides[key] = val
                if overrides:
                    pending_overrides[fp] = overrides
                    imported += 1

            wb.close()
            if added_files:
                w = self.window()
                if w and hasattr(w, "file_tab"):
                    w.file_tab.add_files(added_files)
                else:
                    self._file_paths.extend(added_files)
            self._overrides.update(pending_overrides)
            self._rebuild_table()
            QMessageBox.information(
                self, "Import Complete",
                f"Imported parameters for {imported} file(s)."
            )
        except Exception as exc:
            QMessageBox.warning(self, "Import Error", str(exc))

    def _user_key_from_header(self, header):
        safe = "".join(ch.lower() if ch.isalnum() else "_" for ch in header.strip())
        safe = "_".join(part for part in safe.split("_") if part)
        return CUSTOM_PREFIX + (safe or "field")

    def _resolve_excel_source(self, value, excel_dir):
        raw = Path(value)
        candidates = [raw if raw.is_absolute() else (excel_dir / raw)]
        candidates.extend(Path(fp) for fp in self._file_paths if Path(fp).name == raw.name)
        for candidate in candidates:
            try:
                resolved = candidate.resolve()
            except OSError:
                continue
            if resolved.is_file():
                return resolved
        return None

    def _add_custom_widget_for_key(self, key, label):
        edit = QLineEdit()
        edit.setPlaceholderText(label)
        remove_btn = QPushButton("✕")
        remove_btn.setFixedWidth(28)
        remove_btn.setToolTip(f"Remove '{label}'")
        remove_btn.clicked.connect(lambda: self._remove_custom_field_by_key(key))
        row_layout = QHBoxLayout()
        row_layout.addWidget(edit, 1)
        row_layout.addWidget(remove_btn)
        self._custom_container.addLayout(row_layout)
        self._custom_widgets[key] = (edit, remove_btn, row_layout)

    # ================================================================
    #  Excel template export
    # ================================================================

    def _export_template(self):
        path, _ = QFileDialog.getSaveFileName(
            self, "Export Excel Template", "parameters_template.xlsx",
            "Excel Files (*.xlsx)"
        )
        if not path:
            return
        try:
            import openpyxl
        except ImportError:
            QMessageBox.warning(
                self, "Missing Package",
                "The 'openpyxl' package is required.\n"
                "Install it with:  pip install openpyxl"
            )
            return

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Parameters"

        # Get headers in visual order
        hdr = self.table.horizontalHeader()
        n_cols = self.table.columnCount()
        vis_order = sorted(range(n_cols), key=lambda c: hdr.visualIndex(c))

        header_row = []
        header_row.append("file")
        for logical_col in vis_order:
            if logical_col == 0:
                continue
            if self.table.isColumnHidden(logical_col):
                continue
            header_row.append(self._all_field_keys[logical_col - 1])
        ws.append(header_row)

        # example row
        example = ["sample_data_001.txt"]
        for logical_col in vis_order:
            if logical_col == 0:
                continue
            if self.table.isColumnHidden(logical_col):
                continue
            key = self._all_field_keys[logical_col - 1]
            # provide hint value for some known keys
            hints = {
                "sample_name": "MySample", "sample_id": "001",
                "photon_energy_eV": "21.2", "work_function_eV": "4.2",
                "temperature_K": "300", "polarization": "p-pol",
            }
            example.append(hints.get(key, ""))
        ws.append(example)

        # set column widths
        for i, _ in enumerate(header_row, start=1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 16

        wb.save(path)
        QMessageBox.information(
            self, "Template Exported",
            f"Template saved to:\n{path}\n\n"
            "Fill in parameters for each file and use 'Import from Excel' to load."
        )

    # ================================================================
    #  navigation
    # ================================================================

    def _go_back(self):
        w = self.window()
        if w and hasattr(w, "tabs"):
            w.tabs.setCurrentIndex(0)

    def _go_next(self):
        w = self.window()
        if w and hasattr(w, "tabs"):
            w.tabs.setCurrentIndex(2)
